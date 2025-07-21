import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows
from io import BytesIO
import re
from datetime import datetime
from openpyxl.utils import get_column_letter

class PivotProcessor:
    def process(self, forecast_files, order_file, sales_file, mapping_file):
        from mapping_utils import apply_mapping_and_merge, apply_extended_substitute_mapping, split_mapping_data
        from info_extract import extract_all_year_months, fill_order_data, fill_sales_data, highlight_by_detecting_column_headers
        from name_utils import build_main_df
        from forecast_utils import load_forecast_files, reorder_columns_by_month, merge_monthly_group_headers, merge_and_color_monthly_group_headers, drop_order_shipping_without_forecast
        


        # ✅ 加载原始预测文件
        forecast_dfs = load_forecast_files(forecast_files)
        mapping_semi, mapping_new, mapping_sub = split_mapping_data(mapping_file)
        main_df = build_main_df(forecast_dfs, order_file, sales_file, mapping_new, mapping_sub)

        FIELD_MAPPINGS = {
            "forecast": {"品名": "生产料号"},
            "order": {"品名": "品名"},
            "sales": {"品名": "品名"}
        }

        # ✅ 替换预测中品名
        def apply_mapping_to_all_forecasts(forecast_dfs: dict[str, pd.DataFrame], mapping_new, mapping_sub) -> dict[str, pd.DataFrame]:
            from mapping_utils import apply_mapping_and_merge, apply_extended_substitute_mapping
            mapped_dfs = {}
            for name, df in forecast_dfs.items():
                if df.shape[1] < 2:
                    continue
                second_col = df.columns[1]
                field_mapping = {"品名": second_col}
                try:
                    df_mapped, _ = apply_mapping_and_merge(df.copy(), mapping_new, field_mapping)
                    df_mapped, _ = apply_extended_substitute_mapping(df_mapped, mapping_sub, field_mapping)
                    mapped_dfs[name] = df_mapped
                except KeyError as e:
                    raise ValueError(f"❌ `{name}` 缺失列: {e}. 实际列: {df.columns.tolist()}") from e
            return mapped_dfs

        forecast_dfs = apply_mapping_to_all_forecasts(forecast_dfs, mapping_new, mapping_sub)
        order_file, _ = apply_mapping_and_merge(order_file, mapping_new, FIELD_MAPPINGS["order"])
        order_file, _ = apply_extended_substitute_mapping(order_file, mapping_sub, FIELD_MAPPINGS["order"])
        sales_file, _ = apply_mapping_and_merge(sales_file, mapping_new, FIELD_MAPPINGS["sales"])
        sales_file, _ = apply_extended_substitute_mapping(sales_file, mapping_sub, FIELD_MAPPINGS["sales"])

        # ✅ 提取所有月份（订单/出货用）
        all_months = extract_all_year_months(forecast_dfs, order_file, sales_file)
        for ym in all_months:
            main_df[f"{ym}-订单"] = 0
            main_df[f"{ym}-出货"] = 0

        # ✅ 填充所有预测数据（独立列）
        def extract_file_date(file_name: str) -> str:
            match = re.search(r"(\d{8})", file_name)
            return match.group(1) if match else "00000000"

        def detect_header_row(df: pd.DataFrame) -> int:
            for i, row in df.iterrows():
                if any(isinstance(cell, str) and "产品型号" in str(cell) for cell in row):
                    return i
            return 0

        def standardize_column_name(forecast_col: str, file_date: str) -> str:
            """
            将原始预测列名（如“6月预测”）标准化为“yyyy-mm的预测（yyyy-mm生成）”，处理跨年。
            """
            month_match = re.match(r"^(\d{1,2})月预测$", forecast_col.strip())
            alt_match = re.match(r"^(\d{1,2})月预测\d*$", forecast_col.strip())
            if month_match or alt_match:
                forecast_month = int((month_match or alt_match).group(1))
            else:
                return f"{file_date}-{forecast_col.strip()}"  # fallback: 原样列名
        
            file_year = int(file_date[:4])
            file_month = int(file_date[4:6])
            
            # ✅ 处理跨年：如果预测月份小于生成月份，则年份加一
            if forecast_month < file_month:
                forecast_year = file_year + 1
            else:
                forecast_year = file_year
        
            forecast_month_str = str(forecast_month).zfill(2)
            file_month_str = str(file_month).zfill(2)
            return f"{forecast_year}-{forecast_month_str}的预测（{file_year}-{file_month_str}生成）"


        def fill_forecast_data(main_df: pd.DataFrame, forecast_dfs: dict[str, pd.DataFrame]) -> pd.DataFrame:
            for file_name, df in forecast_dfs.items():
                file_date = extract_file_date(file_name)
                name_col = "生产料号" if "生产料号" in df.columns else (df.columns[1] if df.shape[1] >= 2 else None)
                if name_col is None:
                    continue
                df["品名"] = df[name_col].astype(str).str.strip()
                for col in df.columns:
                    if isinstance(col, str) and "预测" in col:
                        new_col = standardize_column_name(col, file_date)
                        if new_col not in main_df.columns:
                            main_df[new_col] = 0
                        forecast_series = df[["品名", col]].dropna().groupby("品名")[col].sum(min_count=1)
                        main_df[new_col] = main_df["品名"].map(forecast_series).fillna(0)
            return main_df

        main_df = fill_forecast_data(main_df, forecast_dfs)
        main_df = fill_order_data(main_df, order_file, all_months)
        main_df = fill_sales_data(main_df, sales_file, all_months)

        main_df = reorder_columns_by_month(main_df)
        main_df = drop_order_shipping_without_forecast(main_df)

        # 删除所有数值列（除前3列）都为 0 或空的行
        value_cols = main_df.columns[3:]  # 假设前三列为识别字段
        main_df = main_df[~(main_df[value_cols].fillna(0) == 0).all(axis=1)]

        st.write(main_df)

        # ✅ 写入 Excel
        output = BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            main_df.to_excel(writer, index=False, sheet_name="预测分析", startrow=1)
            ws = writer.sheets["预测分析"]
            merge_monthly_group_headers(ws, main_df)
            merge_and_color_monthly_group_headers(ws, main_df)

            for col_idx, column_cells in enumerate(ws.columns, 1):
                max_length = 0
                for cell in column_cells:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                ws.column_dimensions[get_column_letter(col_idx)].width = max_length + 10

            # ✅ 生成“月度预测明细”长表结构
            long_records = []
            pattern = re.compile(r"(\d{4}-\d{2})的预测（(\d{4}-\d{2})生成）")
            for col in main_df.columns:
                match = pattern.match(col)
                if match:
                    forecast_month, file_month = match.groups()
                    for idx, row in main_df.iterrows():
                        long_records.append({
                            "品名": row["品名"],
                            "月份": forecast_month,
                            "生成时间": file_month,
                            "类型": "预测",
                            "数值": row[col]
                        })
                elif col.endswith("订单") or col.endswith("出货"):
                    ym = col[:-2]
                    kind = col[-2:]
                    for idx, row in main_df.iterrows():
                        long_records.append({
                            "品名": row["品名"],
                            "月份": ym,
                            "生成时间": "",  # 订单/出货没有生成时间
                            "类型": kind,
                            "数值": row[col]
                        })

            df_long = pd.DataFrame(long_records)
            df_long = df_long[["品名", "月份", "生成时间", "类型", "数值"]]  # 控制列顺序
            df_long.to_excel(writer, index=False, sheet_name="预测明细")

        output.seek(0)
        return main_df, output
