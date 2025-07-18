import re
import pandas as pd
import streamlit as st
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


import pandas as pd
import re
from datetime import datetime

def extract_all_year_months(forecast_dfs: dict[str, pd.DataFrame], df_order, df_sales, forecast_year=None) -> list[str]:
    if forecast_year is None:
        forecast_year = datetime.today().year

    month_pattern = re.compile(r"(\d{1,2})月预测")
    forecast_months = []

    # 1. 遍历所有预测表提取 “x月预测” 列
    forecast_dfs.values():
        for col in df.columns:
            match = month_pattern.match(str(col).strip())
            if match:
                forecast_m = int(match.group(1))
                # ✅ 跨年判断：预测月份 < 文件生成月份 → 跨年
                if forecast_m < file_month:
                    year = file_year + 1
                else:
                    year = file_year
                month = str(forecast_m).zfill(2)
                forecast_months.append(f"{year}-{month}")

    # 2. 提取 order 文件第 12 列的月份（假设为“订单日期”）
    try:
        order_date_col = df_order.columns[11]
        df_order[order_date_col] = pd.to_datetime(df_order[order_date_col], format="%Y-%m", errors="coerce")
        order_months = (
            df_order[order_date_col]
            .dropna()
            .dt.to_period("M")
            .astype(str)
            .unique()
            .tolist()
        )
    except Exception:
        order_months = []

    # 3. 提取 sales 文件第 6 列的月份（假设为“交易日期”）
    try:
        sales_date_col = df_sales.columns[5]
        df_sales[sales_date_col] = pd.to_datetime(df_sales[sales_date_col], format="%Y-%m", errors="coerce")
        sales_months = (
            df_sales[sales_date_col]
            .dropna()
            .dt.to_period("M")
            .astype(str)
            .unique()
            .tolist()
        )
    except Exception:
        sales_months = []

    # 4. 合并所有月份来源并去重
    all_months = sorted(set(forecast_months + order_months + sales_months))

    # 5. 生成从最小到最大之间的所有月份
    if all_months:
        min_month = pd.Period(min(all_months), freq="M")
        max_month = pd.Period(max(all_months), freq="M")
        full_months = [str(p) for p in pd.period_range(min_month, max_month, freq="M")]
    else:
        full_months = []

    return full_months

def fill_forecast_data(main_df: pd.DataFrame, forecast_dfs: dict[str, pd.DataFrame]) -> pd.DataFrame:
    """
    从 forecast_dfs 中提取所有“x月的预测”列，按品名写入 main_df 中。
    不合并同月不同文件的预测，而是生成多个独立列。
    """
    month_pattern = re.compile(r"^(\d{4})[-年](\d{1,2})月的预测$")

    for file_name, df in forecast_dfs.items():
        if df.shape[1] < 2:
            continue

        # 第2列作为品名列
        name_col = df.columns[1]
        df["品名"] = df[name_col].astype(str).str.strip()

        for col in df.columns:
            match = month_pattern.match(str(col).strip())
            if not match:
                continue

            # 构造唯一列名：文件名-列名
            clean_col_name = f"{file_name}-{col}".replace(".xlsx", "").replace(".xls", "").strip()
            if clean_col_name not in main_df.columns:
                main_df[clean_col_name] = 0

            # 提取并映射
            forecast_series = df[["品名", col]].dropna().groupby("品名")[col].sum(min_count=1)
            main_df[clean_col_name] = main_df["品名"].map(forecast_series).fillna(0)

    return main_df




def fill_order_data(main_df, df_order, forecast_months):
    """
    将订单数据按“订单日期”和“品名”聚合并填入 main_df 中每月的“订单”列。
    
    参数：
    - main_df: 主计划 DataFrame，需包含“品名”列
    - df_order: 上传的未交订单 DataFrame，包含“订单日期”和“品名”
    - forecast_months: 所有涉及的 yyyy-mm 字符串列表
    """
    df_order = df_order.copy()

    # 确保日期字段为 datetime 类型
    df_order["客户要求交期"] = pd.to_datetime(df_order["客户要求交期"], errors="coerce")
    df_order["年月"] = df_order["客户要求交期"].dt.to_period("M").astype(str)

    # 数值字段清洗
    df_order["订单数量"] = pd.to_numeric(df_order["订单数量"], errors="coerce").fillna(0)

    # 聚合出每品名每月的订单量
    grouped = (
        df_order.groupby(["品名", "年月"])["订单数量"]
        .sum()
        .unstack()
        .fillna(0)
    )

    for ym in forecast_months:
        colname = f"{ym}-订单"
        if colname in main_df.columns and ym in grouped.columns:
            main_df[colname] = main_df["品名"].map(grouped[ym]).fillna(0)

    return main_df

def fill_sales_data(main_df, df_sales, forecast_months):
    """
    将出货数据按“交易日期”和“品名”聚合并填入 main_df 中每月的“出货”列。
    
    参数：
    - main_df: 主计划 DataFrame，需包含“品名”列
    - df_sales: 出货明细 DataFrame，包含“交易日期”和“品名”
    - forecast_months: 所有涉及的 yyyy-mm 字符串列表
    """
    df_sales = df_sales.copy()

    # 确保交易日期为 datetime
    df_sales["交易日期"] = pd.to_datetime(df_sales["交易日期"], errors="coerce")
    df_sales["年月"] = df_sales["交易日期"].dt.to_period("M").astype(str)

    # 数值字段清洗
    df_sales["数量"] = pd.to_numeric(df_sales["数量"], errors="coerce").fillna(0)

    # 聚合：每品名每月出货数量
    grouped = (
        df_sales.groupby(["品名", "年月"])["数量"]
        .sum()
        .unstack()
        .fillna(0)
    )

    for ym in forecast_months:
        colname = f"{ym}-出货"
        if colname in main_df.columns and ym in grouped.columns:
            main_df[colname] = main_df["品名"].map(grouped[ym]).fillna(0)

    return main_df

def highlight_by_detecting_column_headers(ws):
    """
    自动识别表头第二行中连续的“预测/订单”列对，并对值为：预测>0且订单=0 的单元格标红。
    """
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    max_col = ws.max_column
    max_row = ws.max_row

    header = [cell.value for cell in ws[2]]

    # 遍历所有列找出“预测-订单”成对列索引
    column_pairs = []
    for i in range(len(header) - 1):
        name1 = str(header[i]).strip()
        name2 = str(header[i + 1]).strip()
        if name1.endswith("预测") and name2.endswith("订单"):
            column_pairs.append((i + 1, i + 2))  # openpyxl列从1开始

    # 遍历每行，检查成对列
    for row in range(3, ws.max_row + 1):
        for forecast_col, order_col in column_pairs:
            cell_forecast = ws.cell(row=row, column=forecast_col)
            cell_order = ws.cell(row=row, column=order_col)

            try:
                val_forecast = float(cell_forecast.value or 0)
                val_order = float(cell_order.value or 0)
            except:
                continue

            if val_forecast > 0 and val_order == 0:
                cell_forecast.fill = red_fill
                cell_order.fill = red_fill
