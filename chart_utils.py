import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows
from io import BytesIO
import re
from datetime import datetime

def write_grouped_forecast_sheet(wb, df: pd.DataFrame, sheet_name="预测展示"):

    from openpyxl.worksheet.worksheet import Worksheet
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    ws: Worksheet = wb.create_sheet(title=sheet_name)

    fixed_cols = ["品名", "月份"]  # 你的前两列字段
    forecast_cols = [col for col in df.columns if "预测" in col and "生成" in col]
    
    # 提取生成时间作为分组依据
    grouped = {}
    for col in forecast_cols:
        match = re.search(r"（(.*?)生成）", col)
        if match:
            gen_month = match.group(1)
            grouped.setdefault(gen_month, []).append(col)

    # 按生成月份排序（yyyy-mm）
    grouped = dict(sorted(grouped.items()))

    # 构造表头
    col_headers = fixed_cols[:]
    sub_headers = fixed_cols[:]
    merge_ranges = []

    for gen_month, cols in grouped.items():
        # 按月份从小到大排序（yyyy-mm）
        cols = sorted(cols)
        grouped[gen_month] = cols
        col_headers.extend([f"{gen_month}生成"] * len(cols))
        sub_headers.extend([re.search(r"^(.*?)的预测", col).group(1) for col in cols])

    # 写入第一行（合并前的内容）
    ws.append(col_headers)
    ws.append(sub_headers)

    # 合并生成月份单元格
    col_idx = len(fixed_cols) + 1
    for gen_month, cols in grouped.items():
        start = col_idx
        end = col_idx + len(cols) - 1
        if start != end:
            ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=end)
        else:
            cell = ws.cell(row=1, column=start)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=1, column=start).value = f"{gen_month}生成"
        col_idx += len(cols)

    # 居中 + 加粗固定列
    for i in range(1, len(fixed_cols) + 1):
        for row in [1, 2]:
            cell = ws.cell(row=row, column=i)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(bold=True)

    # 写入数据行
    for _, row in df.iterrows():
        row_data = [row.get(col, "") for col in fixed_cols]
        for gen_month in grouped:
            for forecast_col in grouped[gen_month]:
                row_data.append(row.get(forecast_col, ""))
        ws.append(row_data)

    # 设置列宽
    for i, col_cells in enumerate(ws.columns, 1):
        max_len = max((len(str(cell.value)) if cell.value else 0) for cell in col_cells)
        ws.column_dimensions[get_column_letter(i)].width = max_len + 5



def write_forecast_expanded_sheet(wb, df: pd.DataFrame, sheet_name="预测展开"):
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.styles import Alignment, Font
    import re

    forecast_cols = [col for col in df.columns if "预测" in col and "生成" in col]
    id_cols = ["品名"]

    records = []
    for _, row in df.iterrows():
        base = {col: row[col] for col in id_cols}
        for col in forecast_cols:
            match = re.match(r"(\d{4}-\d{2})的预测（(\d{4}-\d{2})生成）", col)
            if not match:
                continue
            forecast_month, generated_month = match.groups()
            forecast_value = row[col]
            records.append({
                "品名": base["品名"],
                "预测月份": forecast_month,
                "生成月份": generated_month,
                "预测值": forecast_value,
                "订单量": row.get(f"{forecast_month}-订单", None),
                "出货量": row.get(f"{forecast_month}-出货", None)
            })

    df_out = pd.DataFrame(records)
    df_out = df_out[["品名", "预测月份", "生成月份", "预测值", "订单量", "出货量"]]

    ws = wb.create_sheet(title=sheet_name)
    for r in dataframe_to_rows(df_out, index=False, header=True):
        ws.append(r)

    for cell in ws[1]:
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(bold=True)

    from openpyxl.utils import get_column_letter
    for i, col_cells in enumerate(ws.columns, 1):
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col_cells)
        ws.column_dimensions[get_column_letter(i)].width = max_len + 4

