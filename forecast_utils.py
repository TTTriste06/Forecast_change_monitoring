import pandas as pd
import streamlit as st
import re
from datetime import datetime
from io import BytesIO

from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Alignment, Font

from openpyxl.styles import PatternFill

def drop_order_shipping_without_forecast(main_df: pd.DataFrame) -> pd.DataFrame:
    """
    删除那些没有对应预测列的月份的“订单”列和“出货”列。
    """
    # 提取所有有预测的月份
    forecast_pattern = re.compile(r"^(\d{4}-\d{2})的预测")
    forecast_months = set()

    for col in main_df.columns:
        match = forecast_pattern.match(str(col))
        if match:
            forecast_months.add(match.group(1))

    # 找出需要删除的“订单”和“出货”列
    drop_cols = []
    order_ship_pattern = re.compile(r"^(\d{4}-\d{2})-(订单|出货)$")

    for col in main_df.columns:
        match = order_ship_pattern.match(str(col))
        if match:
            month = match.group(1)
            if month not in forecast_months:
                drop_cols.append(col)

    return main_df.drop(columns=drop_cols)


def merge_and_color_monthly_group_headers(ws: Worksheet, df: pd.DataFrame, start_row: int = 1):
    """
    合并并着色同一个月份的字段（如“预测/订单/出货”）标题行。
    """
    pattern = re.compile(r"(\d{4}-\d{2})")
    col_groups = {}  # {月份: [列索引]}
    
    for idx, col in enumerate(df.columns, start=1):  # openpyxl 列从 1 开始
        match = pattern.search(str(col))
        if match:
            month = match.group(1)
            col_groups.setdefault(month, []).append(idx)

    fill_colors = [
        "FFF2CC", "D9EAD3", "D0E0E3", "F4CCCC", "EAD1DC", "CFE2F3", "FFE599", "E6B8AF"
    ]

    for i, (month, col_indexes) in enumerate(sorted(col_groups.items())):
        start_col = col_indexes[0]
        end_col = col_indexes[-1]
        cell = ws.cell(row=start_row, column=start_col)
        ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=end_col)
        cell.value = month
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(bold=True)

        fill = PatternFill(start_color=fill_colors[i % len(fill_colors)],
                           end_color=fill_colors[i % len(fill_colors)],
                           fill_type="solid")
        for col in col_indexes:
            ws.cell(row=start_row, column=col).fill = fill
            ws.cell(row=start_row + 1, column=col).fill = fill


def merge_monthly_group_headers(ws: Worksheet, df: pd.DataFrame, start_row: int = 1):
    """
    将同一月份的“预测/订单/出货”等字段在 Excel 中合并单元格并写入“yyyy-mm”。
    
    参数：
        ws: openpyxl 的 worksheet 对象
        df: DataFrame，用于获取列顺序
        start_row: 起始行（默认为 1）
    """
    pattern = re.compile(r"(\d{4}-\d{2})")
    col_groups = {}  # {月份: [列索引]}
    
    for idx, col in enumerate(df.columns, start=1):  # openpyxl 列从 1 开始
        match = pattern.search(str(col))
        if match:
            month = match.group(1)
            col_groups.setdefault(month, []).append(idx)

    for month, col_indexes in col_groups.items():
        if len(col_indexes) >= 2:
            start_col = col_indexes[0]
            end_col = col_indexes[-1]
            cell = ws.cell(row=start_row, column=start_col)
            ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=end_col)
            cell.value = month
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(bold=True)


# ✅ 对列进行排序：按月份分组排序，预测/订单/出货顺序
def reorder_columns_by_month(main_df: pd.DataFrame) -> pd.DataFrame:
    fixed_cols = ["晶圆品名", "规格", "品名"]
    pattern = re.compile(r"(\d{4}-\d{2})")

    # 收集所有字段按年月归类
    month_field_map = {}
    for col in main_df.columns:
        match = pattern.search(str(col))
        if match:
            ym = match.group(1)
            month_field_map.setdefault(ym, []).append(col)

    # 对每月的列再按关键词排序：预测 < 订单 < 出货
    def sort_key(col):
        if "预测" in col:
            return 0
        elif "订单" in col:
            return 1
        elif "出货" in col:
            return 2
        else:
            return 3

    sorted_month_cols = []
    for ym in sorted(month_field_map.keys()):
        sorted_cols = sorted(month_field_map[ym], key=sort_key)
        sorted_month_cols.extend(sorted_cols)

    final_cols = fixed_cols + sorted_month_cols
    return main_df[[col for col in final_cols if col in main_df.columns]]


def extract_forecast_generation_date(self, filename: str) -> str:
    """从文件名中提取生成日期，返回格式为 yyyy-mm"""
    match = re.search(r"_(\d{8})", filename)
    if match:
        date_str = match.group(1)
        try:
            dt = datetime.strptime(date_str, "%Y%m%d")
            return dt.strftime("%Y-%m")
        except ValueError:
            pass
    return "unknown"

def extract_forecast_data(self, file: BytesIO) -> tuple[pd.DataFrame, str]:
    """读取预测文件最长的sheet，并自动识别header行（包含‘产品型号’），返回DataFrame和生成年月"""
    xls = pd.ExcelFile(file)
    max_len = 0
    selected_sheet = None

    for sheet in xls.sheet_names:
        df = xls.parse(sheet, header=None)
        if df.shape[0] > max_len:
            max_len = df.shape[0]
            selected_sheet = sheet

    df_raw = xls.parse(selected_sheet, header=None)

    header_row_idx = None
    for idx, row in df_raw.iterrows():
        if row.astype(str).str.contains("产品型号").any():
            header_row_idx = idx
            break

    if header_row_idx is None:
        raise ValueError("❌ 未在文件中识别到包含‘产品型号’的header行")

    df = pd.read_excel(file, sheet_name=selected_sheet, header=header_row_idx)
    df.columns.values[1] = "品名"
    return df, selected_sheet  # sheet名可选作备份信息

def parse_forecast_months(self, forecast_df: pd.DataFrame, base_year: int) -> dict:
    """
    输入包含“x月预测”列的df，返回一个dict：
    {"yyyy-mm": 原始列名}，自动判断跨年
    """
    pattern = re.compile(r"^(\d{1,2})月预测$")
    col_map = {}

    start_year = base_year
    last_month = 0
    for col in forecast_df.columns:
        match = pattern.match(col)
        if match:
            month = int(match.group(1))
            if last_month and month < last_month:
                start_year += 1
            last_month = month
            ym = f"{start_year}-{month:02d}"
            col_map[ym] = col
    return col_map

def append_multi_forecast_columns(
    self,
    main_df: pd.DataFrame,
    forecast_df: pd.DataFrame,
    col_map: dict,
    label: str
) -> pd.DataFrame:
    """
    在 main_df 中添加来自 forecast_df 的预测列，列名为 “label（yyyy-mm）”，
    forecast_df 中“生产料号”为品名，col_map 为 {yyyy-mm: 原始列名}
    """
    forecast_df["生产料号"] = forecast_df["生产料号"].astype(str).str.strip()
    forecast_df = forecast_df.rename(columns={"生产料号": "品名"})
    main_df["品名"] = main_df["品名"].astype(str).str.strip()

    for ym, orig_col in col_map.items():
        new_col = f"{label}（{ym}）"
        if new_col not in main_df.columns:
            main_df[new_col] = 0
        for i, row in main_df.iterrows():
            name = row["品名"]
            val = forecast_df.loc[forecast_df["品名"] == name, orig_col]
            if not val.empty:
                main_df.at[i, new_col] = val.values[0]
    return main_df


def merge_forecast_columns(forecast_dfs: dict[str, pd.DataFrame]) -> pd.DataFrame:
    """
    合并多个预测文件，按“品名”横向合并所有预测列。
    返回值为：包含“品名”列 + 多个“yyyy-mm的预测（生成年月）”列
    """
    all_parts = []
    for file_name, df in forecast_dfs.items():
        try:
            forecast_parts = parse_forecast_columns(df, file_name)
            for subdf in forecast_parts.values():
                all_parts.append(subdf)
        except Exception as e:
            st.warning(f"⚠ 处理 {file_name} 失败：{e}")

    # 统一合并
    if not all_parts:
        return pd.DataFrame(columns=["品名"])
    merged = all_parts[0]
    for part in all_parts[1:]:
        merged = pd.merge(merged, part, on="品名", how="outer")

    return merged.fillna(0)


def parse_forecast_columns(df: pd.DataFrame, file_name: str) -> dict[str, pd.Series]:
    """
    从预测表中提取“x月预测”列，并转换为“yyyy-mm的预测（生成年月）”格式
    """
    # 从文件名中提取生成日期
    match = re.search(r"(\d{8})", file_name)
    if not match:
        raise ValueError(f"文件名中未找到 8 位日期：{file_name}")
    gen_date = pd.to_datetime(match.group(1), format="%Y%m%d")
    gen_year = gen_date.year
    gen_month = gen_date.month
    gen_ym_str = gen_date.strftime("%Y-%m")

    forecast_cols = {}
    month_pattern = re.compile(r"^(\d{1,2})月预测$")

    for col in df.columns:
        if isinstance(col, str):
            match = month_pattern.match(col.strip())
            if match:
                month_num = int(match.group(1))
                # 决定年份（跨年判断）
                if month_num >= gen_month:
                    year = gen_year
                else:
                    year = gen_year + 1
                ym_key = f"{year}-{str(month_num).zfill(2)}"
                new_col_name = f"{ym_key}的预测（{gen_ym_str}）"
                forecast_cols[new_col_name] = df[[col, "品名"]].rename(columns={col: new_col_name})

    return forecast_cols  # 返回列名 → dataframe with 品名 + 单列

def load_forecast_files(files: dict) -> dict[str, pd.DataFrame]:
    """
    对上传的多个预测 Excel 文件执行以下操作：
    - 找到每个文件中最长的 sheet
    - 自动识别 header 行（含“产品型号”的那一行）
    - 将第二列统一命名为“品名”
    - 用 st.write 打印每个文件读取结果
    返回值：dict[file_name -> cleaned DataFrame]
    """
    result = {}

   

    for uploaded_file in files:
        file_name = uploaded_file.name
        try:
            xls = pd.ExcelFile(uploaded_file)
            longest_sheet = max(xls.sheet_names, key=lambda name: pd.read_excel(xls, sheet_name=name).shape[0])
            df_raw = pd.read_excel(xls, sheet_name=longest_sheet, header=None)

            # 自动识别 header 行：包含“产品型号”的行
            header_row_idx = df_raw[df_raw.apply(lambda row: row.astype(str).str.contains("产品型号").any(), axis=1)].index
            if header_row_idx.empty:
                st.warning(f"⚠ 文件 {file_name} 中未找到包含“产品型号”的表头行，跳过")
                continue

            header_row = header_row_idx[0]
            df = pd.read_excel(xls, sheet_name=longest_sheet, header=header_row)

            # 统一第二列为“品名”
            if df.shape[1] >= 2:
                df.columns.values[1] = "品名"

            st.write(f"📄 读取成功：{file_name}（使用 sheet：{longest_sheet}，header 行：第 {header_row+1} 行）")
            st.dataframe(df)

            result[file_name] = df

        except Exception as e:
            st.error(f"❌ 无法读取文件 {file_name}: {e}")

    return result
